from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
import uuid
import io
import openpyxl

from ..database import get_db
from ..models import StaffAssignment, Subject, User, UserRole
from ..schemas import (
    AssignStaffRequest, UpdatePermissionsRequest,
    StaffAssignmentResponse, MySubjectAssignment, FacultyListResponse,
    StaffImportResult, AllStaffResponse,
)
from ..schemas.user import UserCreate
from ..services.auth import get_current_user, hash_password

router = APIRouter(prefix="/staff", tags=["Staff Management"])


# ── Faculty / staff listing ────────────────────────────────────────────────────

@router.get("/faculty-list", response_model=List[FacultyListResponse])
async def get_faculty_list(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Get all faculty members from the database"""
    faculty = db.query(User).filter(User.role.in_([UserRole.FACULTY, UserRole.HOD])).all()
    return [FacultyListResponse(
        id=f.id,
        email=f.email,
        name=f.name,
        department=f.department
    ) for f in faculty]


@router.get("/all", response_model=List[AllStaffResponse])
async def get_all_staff(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """HOD-only: list every registered staff/faculty member with their subject assignments."""
    if user.role != UserRole.HOD:
        raise HTTPException(status_code=403, detail="Only HOD can view all staff")

    faculty = db.query(User).filter(User.role == UserRole.FACULTY).all()

    result = []
    for f in faculty:
        # Collect subject codes for this staff member
        assignments = db.query(StaffAssignment).filter(
            StaffAssignment.staff_email == f.email,
            StaffAssignment.is_active == True
        ).all()
        subject_codes = []
        for a in assignments:
            subject = db.query(Subject).filter(Subject.id == a.subject_id).first()
            if subject:
                subject_codes.append(subject.code)

        result.append(AllStaffResponse(
            id=f.id,
            email=f.email,
            name=f.name,
            department=f.department,
            is_active=f.is_active,
            assigned_subjects=subject_codes,
        ))

    return result


# ── Subject assignments ────────────────────────────────────────────────────────

@router.get("/my-subjects", response_model=List[MySubjectAssignment])
async def get_my_subjects(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Get subjects assigned to current staff with full configuration"""
    assignments = db.query(StaffAssignment).filter(
        StaffAssignment.staff_email == user.email,
        StaffAssignment.is_active == True
    ).all()
    
    result = []
    for a in assignments:
        subject = db.query(Subject).filter(Subject.id == a.subject_id).first()
        if subject:
            result.append(MySubjectAssignment(
                subjectId=a.subject_id,
                subjectName=subject.name,
                subjectCode=subject.code,
                subjectNature=subject.nature,
                subjectConfiguration=subject.configuration,
                canEditPattern=a.can_edit_pattern,
                canGenerateQuestions=a.can_generate_questions,
                canApprove=a.can_approve
            ))
    
    return result


@router.get("/subject/{subject_id}", response_model=List[StaffAssignmentResponse])
async def get_subject_staff(
    subject_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Get all staff assigned to a subject"""
    return db.query(StaffAssignment).filter(
        StaffAssignment.subject_id == subject_id,
        StaffAssignment.is_active == True
    ).all()


@router.post("/assign/{subject_id}", response_model=StaffAssignmentResponse)
async def assign_staff(
    subject_id: str,
    data: AssignStaffRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Assign staff to a subject"""
    if user.role != UserRole.HOD:
        raise HTTPException(status_code=403, detail="Only HOD can assign staff")
    
    existing = db.query(StaffAssignment).filter(
        StaffAssignment.subject_id == subject_id,
        StaffAssignment.staff_email == data.staff_email
    ).first()
    
    if existing:
        existing.is_active = True
        existing.can_edit_pattern = data.permissions.canEditPattern
        existing.can_generate_questions = data.permissions.canGenerateQuestions
        existing.can_approve = data.permissions.canApprove
        existing.staff_name = data.staff_name or existing.staff_name
        db.commit()
        db.refresh(existing)
        return existing
    
    assignment = StaffAssignment(
        id=str(uuid.uuid4()),
        subject_id=subject_id,
        staff_email=data.staff_email,
        staff_name=data.staff_name,
        can_edit_pattern=data.permissions.canEditPattern,
        can_generate_questions=data.permissions.canGenerateQuestions,
        can_approve=data.permissions.canApprove,
        assigned_by=user.id
    )
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    return assignment


@router.put("/assignment/{assignment_id}", response_model=StaffAssignmentResponse)
async def update_permissions(
    assignment_id: str,
    data: UpdatePermissionsRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Update staff permissions"""
    if user.role != UserRole.HOD:
        raise HTTPException(status_code=403, detail="Only HOD can update permissions")
    
    assignment = db.query(StaffAssignment).filter(StaffAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    assignment.can_edit_pattern = data.permissions.canEditPattern
    assignment.can_generate_questions = data.permissions.canGenerateQuestions
    assignment.can_approve = data.permissions.canApprove
    
    db.commit()
    db.refresh(assignment)
    return assignment


@router.delete("/assignment/{assignment_id}")
async def remove_staff(
    assignment_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Remove staff from subject"""
    if user.role != UserRole.HOD:
        raise HTTPException(status_code=403, detail="Only HOD can remove staff")
    
    assignment = db.query(StaffAssignment).filter(StaffAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    assignment.is_active = False
    db.commit()
    return {"message": "Staff removed"}


@router.delete("/{staff_id}")
async def delete_staff_member(
    staff_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """HOD-only: permanently deactivate a staff member account."""
    if user.role != UserRole.HOD:
        raise HTTPException(status_code=403, detail="Only HOD can remove staff members")

    staff = db.query(User).filter(User.id == staff_id, User.role == UserRole.FACULTY).first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")

    # Deactivate all assignments
    db.query(StaffAssignment).filter(
        StaffAssignment.staff_email == staff.email
    ).update({"is_active": False})

    # Deactivate user
    staff.is_active = False
    db.commit()
    return {"message": "Staff member deactivated"}


# ── Excel bulk import ──────────────────────────────────────────────────────────

@router.post("/import-excel", response_model=StaffImportResult)
async def import_staff_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    HOD-only: bulk-import staff from an Excel file.

    Expected columns (case-insensitive):
      - name  (or staff_name)
      - email (or staff_email)
      - subjects (optional, comma-separated subject codes e.g. "DBMS, CN, OS")

    If the 'subjects' column is absent entirely, staff identities are created
    without any subject assignment. The HOD can assign subjects manually later.
    """
    if user.role != UserRole.HOD:
        raise HTTPException(status_code=403, detail="Only HOD can import staff")

    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }
    content_type = file.content_type or ""
    filename_lower = (file.filename or "").lower()
    if content_type not in allowed_types and not filename_lower.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Please upload an Excel file (.xlsx or .xls)"
        )

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    # ── Parse headers ──────────────────────────────────────────────────────────
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not headers_row:
        raise HTTPException(status_code=400, detail="Excel file appears to be empty")

    # Normalise header names
    headers = {
        (str(h).strip().lower() if h else ""): idx
        for idx, h in enumerate(headers_row)
    }

    # Map to our canonical column names
    COL_ALIASES = {
        "name":     ["name", "staff_name", "staff name", "faculty name"],
        "email":    ["email", "staff_email", "staff email", "faculty email", "e-mail"],
        "subjects": ["subjects", "subject", "subject_codes", "subject codes", "assigned subjects"],
    }

    col_map: dict = {}
    for field, aliases in COL_ALIASES.items():
        for alias in aliases:
            if alias in headers:
                col_map[field] = headers[alias]
                break

    if "name" not in col_map:
        raise HTTPException(status_code=400, detail="Column 'name' (or 'staff_name') not found in Excel")
    if "email" not in col_map:
        raise HTTPException(status_code=400, detail="Column 'email' (or 'staff_email') not found in Excel")

    has_subjects_col = "subjects" in col_map

    result = StaffImportResult()
    DEFAULT_TEMP_PASSWORD = "Welcome@123"  # Staff must change on first login

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Extract cell values
        def cell(field: str):
            idx = col_map.get(field)
            if idx is None:
                return None
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else None

        name  = cell("name")
        email = cell("email")
        subjects_raw = cell("subjects") if has_subjects_col else None

        if not name and not email:
            continue  # Skip blank rows silently

        if not name:
            result.errors.append(f"Row {row_num}: missing name — skipped")
            continue
        if not email:
            result.errors.append(f"Row {row_num}: missing email — skipped")
            continue

        # Basic email format check
        if "@" not in email or "." not in email.split("@")[-1]:
            result.errors.append(f"Row {row_num}: invalid email '{email}' — skipped")
            continue

        # ── Create or locate user ──────────────────────────────────────────────
        existing_user = db.query(User).filter(User.email == email).first()
        if existing_user:
            result.updated += 1
            staff_user = existing_user
        else:
            staff_user = User(
                id=str(uuid.uuid4()),
                email=email,
                password=hash_password(DEFAULT_TEMP_PASSWORD),
                name=name,
                role=UserRole.FACULTY,
                department=user.department,  # inherit HOD's department
                is_active=True,
            )
            db.add(staff_user)
            try:
                db.flush()  # Get the ID without committing yet
            except Exception as e:
                db.rollback()
                result.errors.append(f"Row {row_num}: DB error creating user — {e}")
                continue
            result.created += 1

        # ── Assign subjects if column present ──────────────────────────────────
        if has_subjects_col and subjects_raw:
            codes = [c.strip() for c in subjects_raw.replace(",", ";").split(";") if c.strip()]
            for code in codes:
                subject = db.query(Subject).filter(Subject.code == code).first()
                if not subject:
                    result.errors.append(
                        f"Row {row_num}: subject code '{code}' not found — assignment skipped"
                    )
                    continue

                # Check if already assigned
                existing_assign = db.query(StaffAssignment).filter(
                    StaffAssignment.subject_id == subject.id,
                    StaffAssignment.staff_email == email,
                ).first()

                if existing_assign:
                    existing_assign.is_active = True
                    existing_assign.staff_name = name
                else:
                    db.add(StaffAssignment(
                        id=str(uuid.uuid4()),
                        subject_id=subject.id,
                        staff_email=email,
                        staff_name=name,
                        can_edit_pattern=False,
                        can_generate_questions=True,
                        can_approve=False,
                        assigned_by=user.id,
                    ))
                    result.assignments_added += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to save import: {e}")

    return result
