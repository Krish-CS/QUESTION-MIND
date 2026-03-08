from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from typing import Dict, List, Any, Optional
import os
import io
import base64
from datetime import datetime

class ExcelService:
    def __init__(self):
        self.upload_dir = "uploads/question-banks"
        os.makedirs(self.upload_dir, exist_ok=True)

    @staticmethod
    def _format_question_text(text: str) -> str:
        """
        Normalise question / answer text so it renders cleanly in Excel cells.
        - Convert escaped \\n sequences to real newlines
        - Collapse 3+ consecutive blank lines to 2
        - Strip leading/trailing whitespace per paragraph
        - Remove stray backslash-n artifacts like '\\n1)' → '\n1)'
        """
        if not text:
            return text
        # Decode escaped newline sequences that survived JSON serialisation
        text = text.replace('\\n', '\n')
        # Remove Windows carriage returns
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        # Collapse 3+ blank lines → max 2
        import re
        text = re.sub(r'\n{3,}', '\n\n', text)
        # Strip trailing spaces on each line
        text = '\n'.join(line.rstrip() for line in text.split('\n'))
        return text.strip()

    def _calculate_row_height(self, text: str, column_width: int, font_size: int = 11) -> float:
        """Calculate appropriate row height based on text content and column width"""
        if not text:
            return 30
        
        # Approximate characters per line based on column width
        # Conservative estimate: width * 0.9 to account for padding and font width variance
        chars_per_line = max(int(column_width * 0.95), 15)
        
        # Count explicit line breaks
        lines = text.split('\n')
        total_lines = 0
        
        for line in lines:
            line_len = len(line)
            if line_len == 0:
                total_lines += 1  # Empty line
            else:
                # Calculate wrapped lines
                # Use math.ceil logic equivalent for wrapping
                wrapped = (line_len + chars_per_line - 1) // chars_per_line
                total_lines += wrapped
        
        # Height per line - increased to 16 for better spacing
        line_height = 16
        calculated_height = total_lines * line_height
        
        # Add generous padding to ensure nothing is cut off
        return max(30, calculated_height + 15)

    def _create_rich_text(self, text: str, normal_font: Font) -> CellRichText:
        """Parse markdown code blocks and return RichText with bold language identifier"""
        if '```' not in text:
            return CellRichText([TextBlock(InlineFont(sz=11), text)])
        
        parts = text.split('```')
        rich_content = []
        
        # Bold font for language identifier
        lang_font = InlineFont(b=True, sz=11)
        # Normal font for code and text
        reg_font = InlineFont(sz=11)
        
        for i, part in enumerate(parts):
            if not part: continue
            
            if i % 2 == 1:
                # Code block
                lines = part.split('\n', 1)
                # Check if first line is a language identifier
                if len(lines) > 0 and lines[0].strip() and lines[0].strip().isalnum():
                    lang_name = lines[0].strip()
                    
                    # Ensure separation from previous text if needed
                    prefix = "\n" if i > 0 else ""
                    
                    # Add language name in BOLD, followed by newline
                    rich_content.append(TextBlock(lang_font, f"{prefix}{lang_name}\n"))
                    
                    # Add the rest of the code in NORMAL font
                    if len(lines) > 1:
                        rich_content.append(TextBlock(reg_font, lines[1]))
                    else:
                        # Empty code block case
                        pass
                else:
                    # No language identifier, just add content
                    # If there's previous text, ensure a newline separation for the code block
                    prefix = "\n" if i > 0 else ""
                    rich_content.append(TextBlock(reg_font, f"{prefix}{part}"))
            else:
                # Regular text
                rich_content.append(TextBlock(reg_font, part))
                
        return CellRichText(rich_content)
    
    def generate_question_bank_excel(
        self,
        questions: Dict[str, List[Dict]],
        subject_name: str,
        subject_code: str,
        parts_config: List[Dict],
        department: str = "CSE",
        semester: int = 1,
        academic_year: str = None,
        has_cdap: bool = False
    ) -> str:
        """Generate Excel file for question bank matching the required format"""
        
        if not academic_year:
            year = datetime.now().year
            academic_year = f"{year}-{year + 1}"
        
        wb = Workbook()
        
        # Common styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # GREY header fill (like in the reference image)
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        info_font = Font(bold=True, size=11)
        normal_font = Font(size=11)
        
        # Column widths - With or without CDAP Part column
        # A: Serial No, B: Question (wide), C: BTL, D: CDAP Part (if has_cdap), then Course Outcomes, Marks, Part
        if has_cdap:
            col_widths_qb = {'A': 8, 'B': 55, 'C': 12, 'D': 12, 'E': 16, 'F': 10, 'G': 10}
            col_widths_ans = {'A': 8, 'B': 45, 'C': 50, 'D': 10, 'E': 8, 'F': 10, 'G': 10}
        else:
            col_widths_qb = {'A': 8, 'B': 60, 'C': 12, 'D': 18, 'E': 12, 'F': 10}
            col_widths_ans = {'A': 8, 'B': 50, 'C': 55, 'D': 8, 'E': 10, 'F': 10}
        
        # ============== SHEET 1: Question Bank ==============
        ws_qb = wb.active
        ws_qb.title = "Question Bank"
        
        # Set column widths first
        for col_letter, width in col_widths_qb.items():
            ws_qb.column_dimensions[col_letter].width = width
        
        # Row 1: Completely BLANK row
        
        # Row 2: Info labels - Start from Column B (Column A is empty)
        # B=Course Code-Course Name, C=Department, D=Semester, E=Academic Year
        ws_qb['B2'] = "Course Code-Course Name"
        ws_qb['C2'] = "Department"
        ws_qb['D2'] = "Semester"
        ws_qb['E2'] = "Academic Year"
        # Make Academic Year column wider to fit the header
        ws_qb.column_dimensions['E'].width = 16
        for col in range(2, 6):
            ws_qb.cell(row=2, column=col).font = info_font
            ws_qb.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 3: Info values - Start from Column B (Column A is empty)
        ws_qb['B3'] = f"{subject_code}- {subject_name}"
        ws_qb['C3'] = department
        ws_qb['D3'] = self._roman_numeral(semester)
        ws_qb['E3'] = academic_year
        for col in range(2, 6):
            ws_qb.cell(row=3, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 4: Empty
        
        # Row 5: Question Bank headers with GREY background
        # Column A is EMPTY (no header), columns B onwards have headers
        if has_cdap:
            qb_headers = ['', 'Question Bank', 'BTL Level', 'CDAP Part', 'Course Outcomes', 'Marks', 'Part']
        else:
            qb_headers = ['', 'Question Bank', 'BTL Level', 'Course Outcomes', 'Marks', 'Part']
        
        for col, header in enumerate(qb_headers, 1):
            cell = ws_qb.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_qb.row_dimensions[5].height = 22
        
        # Row 6+: Questions (NO fill, just borders)
        current_row = 6
        q_number = 1
        
        for part_name, part_questions in questions.items():
            part_config = next((p for p in parts_config if p.get('partName') == part_name), {})
            marks_per_q = part_config.get('marksPerQuestion', 2)
            
            # Sort: CO order (unit 1 → 2 → …), MCQ before descriptive within each CO,
            # then CDAP Part 1 before Part 2 within each group
            sorted_questions = sorted(
                part_questions,
                key=lambda q: (
                    q.get('unit', 1),          # CO1 first, then CO2, CO3 …
                    0 if q.get('isMCQ') else 1, # MCQ before descriptive
                    q.get('cdap_part', 1),      # CDAP Part 1 before Part 2
                )
            )
            
            for q in sorted_questions:
                question_text = self._format_question_text(q.get('question', ''))

                # For MCQ, add options on new lines
                if q.get('isMCQ') and q.get('options'):
                    opts = q['options']
                    question_text += f"\nA) {opts.get('A', '')}"
                    question_text += f"\nB) {opts.get('B', '')}"
                    question_text += f"\nC) {opts.get('C', '')}"
                    question_text += f"\nD) {opts.get('D', '')}"

                unit_num = q.get('unit', 1)
                cdap_part = q.get('cdap_part', 1) if has_cdap else None
                cdap_part_display = f"Part {cdap_part}" if cdap_part else ""
                
                # Build row data based on whether CDAP is present
                if has_cdap:
                    row_data = [
                        q_number,
                        question_text,
                        q.get('btl', 'BTL2'),
                        cdap_part_display,  # CDAP Part column
                        unit_num,
                        q.get('marks', marks_per_q),
                        part_name
                    ]
                else:
                    row_data = [
                        q_number,
                        question_text,
                        q.get('btl', 'BTL2'),
                        unit_num,
                        q.get('marks', marks_per_q),
                        part_name
                    ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws_qb.cell(row=current_row, column=col, value=value)
                    cell.border = thin_border
                    cell.font = normal_font
                    
                    # Q&A Columns: Justified, Vertical Center
                    if col == 2:
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='justify', indent=1)
                    else:
                        # Others: Centered
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                # Calculate and set row height based on question text
                row_height = self._calculate_row_height(question_text, col_widths_qb['B'])

                # Embed image if present
                image_data = q.get('imageData')
                if image_data:
                    try:
                        raw_b64 = image_data.split(',', 1)[1] if ',' in image_data else image_data
                        img_bytes = base64.b64decode(raw_b64)
                        img_buf = io.BytesIO(img_bytes)
                        xl_img = XLImage(img_buf)
                        # Scale down to fit column B (max 200×120 px)
                        max_w, max_h = 200, 120
                        if xl_img.width and xl_img.height:
                            scale = min(max_w / xl_img.width, max_h / xl_img.height, 1.0)
                            xl_img.width = int(xl_img.width * scale)
                            xl_img.height = int(xl_img.height * scale)
                        xl_img.anchor = f'B{current_row}'
                        ws_qb.add_image(xl_img)
                        # Ensure row is tall enough (points ≈ pixels × 0.75)
                        row_height = max(row_height, xl_img.height * 0.75 + 6)
                    except Exception:
                        pass  # Best-effort: skip image if decode fails

                ws_qb.row_dimensions[current_row].height = row_height
                
                current_row += 1
                q_number += 1
        
        # ============== SHEET 2: Answer Key ==============
        ws_ans = wb.create_sheet("Answer Key")
        
        # Set column widths first
        for col_letter, width in col_widths_ans.items():
            ws_ans.column_dimensions[col_letter].width = width
        
        # Title row - span more columns if CDAP is present
        merge_range = 'A1:G1' if has_cdap else 'A1:F1'
        ws_ans.merge_cells(merge_range)
        ws_ans['A1'] = f"{subject_code} - {subject_name}"
        ws_ans['A1'].font = title_font
        ws_ans['A1'].alignment = Alignment(horizontal='center')
        
        # Subtitle
        merge_range = 'A2:G2' if has_cdap else 'A2:F2'
        ws_ans.merge_cells(merge_range)
        ws_ans['A2'] = "ANSWER KEY / EXPECTED ANSWERS"
        ws_ans['A2'].font = Font(bold=True, size=12, color="2E75B6")
        ws_ans['A2'].alignment = Alignment(horizontal='center')
        
        # Row 3: Empty
        
        # Row 4: Headers with GREY background
        if has_cdap:
            ans_headers = ['Sl.No', 'Question', 'Expected Answer', 'CDAP Part', 'BTL', 'Marks', 'Part']
        else:
            ans_headers = ['Sl.No', 'Question', 'Expected Answer', 'BTL', 'Marks', 'Part']
        
        for col, header in enumerate(ans_headers, 1):
            cell = ws_ans.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_ans.row_dimensions[4].height = 22
        
        # Row 5+: Answers
        current_row = 5
        q_number = 1
        
        for part_name, part_questions in questions.items():
            part_config = next((p for p in parts_config if p.get('partName') == part_name), {})
            marks_per_q = part_config.get('marksPerQuestion', 2)
            
            # Sort: CO order (unit 1 → 2 → …), MCQ before descriptive within each CO,
            # then CDAP Part 1 before Part 2 within each group
            sorted_questions = sorted(
                part_questions,
                key=lambda q: (
                    q.get('unit', 1),          # CO1 first, then CO2, CO3 …
                    0 if q.get('isMCQ') else 1, # MCQ before descriptive
                    q.get('cdap_part', 1),      # CDAP Part 1 before Part 2
                )
            )
            
            for q in sorted_questions:
                question_text = self._format_question_text(q.get('question', ''))

                # For MCQ, add options on new lines (same as Question Bank sheet)
                if q.get('isMCQ') and q.get('options'):
                    opts = q['options']
                    question_text += f"\nA) {opts.get('A', '')}"
                    question_text += f"\nB) {opts.get('B', '')}"
                    question_text += f"\nC) {opts.get('C', '')}"
                    question_text += f"\nD) {opts.get('D', '')}"

                # Answer text
                answer_text = self._format_question_text(q.get('answer', ''))
                if q.get('isMCQ') and q.get('correctOption'):
                    opts = q.get('options', {})
                    correct = q['correctOption']
                    answer_text = f"{correct}) {opts.get(correct, answer_text)}"
                
                cdap_part = q.get('cdap_part', 1) if has_cdap else None
                cdap_part_display = f"Part {cdap_part}" if cdap_part else ""
                
                # Build row data based on whether CDAP is present
                if has_cdap:
                    row_data = [
                        q_number,
                        question_text,
                        answer_text,
                        cdap_part_display,  # CDAP Part column
                        q.get('btl', 'BTL2'),
                        q.get('marks', marks_per_q),
                        part_name
                    ]
                else:
                    row_data = [
                        q_number,
                        question_text,
                        answer_text,
                        q.get('btl', 'BTL2'),
                        q.get('marks', marks_per_q),
                        part_name
                    ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws_ans.cell(row=current_row, column=col)
                    cell.border = thin_border
                    
                    # Apply RichText for Answer column (Column 3)
                    if col == 3 and isinstance(value, str) and '```' in value:
                        cell.value = self._create_rich_text(value, normal_font)
                    else:
                        cell.value = value
                        cell.font = normal_font

                    # Alignment configuration
                    if col in [2, 3]: 
                        # Question and Answer: Justified Left, Vertical Center (Middle)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='justify', indent=1)
                    else:
                        # Others: Centered Middle
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                # Calculate row height based on question and answer text
                q_height = self._calculate_row_height(question_text, col_widths_ans['B'])
                a_height = self._calculate_row_height(answer_text, col_widths_ans['C'])
                ws_ans.row_dimensions[current_row].height = max(q_height, a_height)
                
                current_row += 1
                q_number += 1
        
        # Save file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{subject_code}_QuestionBank_{timestamp}.xlsx"
        filepath = os.path.join(self.upload_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def _roman_numeral(self, num: int) -> str:
        """Convert number to Roman numeral"""
        roman_map = {
            1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V',
            6: 'VI', 7: 'VII', 8: 'VIII', 9: 'IX', 10: 'X'
        }
        return roman_map.get(num, str(num))


excel_service = ExcelService()

